import os
import pandas as pd
from openpyxl import load_workbook


def save_data_to_excel(data, file_name="movie_data", file_path="./movie_list/"):
    # File name
    file_name = file_name + ".xlsx"
    file_location = file_path + file_name

    # Create data frame over processed data
    df = pd.DataFrame([data])

    # Check if the file exists
    if os.path.exists(file_location):
        append_data_to_excel(data,file_location)
    else:
        # Store the data in excel files
        create_new_excel(data, file_location)




def create_new_excel(data, file_location):
    # Create data frame over processed data
    df = pd.DataFrame([data])

    # Store the data in a new Excel file
    df.to_excel(file_location, index=False)


def append_data_to_excel(data, file_location):
    # Load the existing Excel file
    book = load_workbook(file_location)
    writer = pd.ExcelWriter(file_location, engine='openpyxl')
    writer.book = book

    # Select the first sheet
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    sheet_name = list(writer.sheets.keys())[0]

    # Append the data to the existing sheet starting from the next empty row
    df = pd.DataFrame([data])
    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=writer.sheets[sheet_name].max_row)

    # Save the changes
    writer.save()
    writer.close()
